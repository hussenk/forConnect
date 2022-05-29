import csv
from flask import request, redirect
from openpyxl import load_workbook


class serves:

    def __init__(self):
        self.arrayData = []

    def stringSpelter(self, str):
        return str.replace(';', ' ').replace(',', ' ').replace('.', ' ').replace('،', ' ').replace('-', ' ').split()

    def handelErrorUpload(self):
        # check method
        # print('check  request method')
        if(request.method == 'GET'):
            return redirect('/')
        # check if file exist
        # print('check if file exist')
        if ('upload' not in request.files):
            return redirect('/')

        self.dColumns = self.stringSpelter(request.form['dColumns'])
        self.newHeaders = self.stringSpelter(request.form['newHeaders'])
        # print(self.dColumns)
        # print(self.newHeaders)
        self.file = request.files['upload']

        # getting file name and extension
        self.name,  self.extension = self.file.filename.rsplit('.', 1)
        # print('getting file name and extension')

        # check type file
        # print('check type file')
        if(self.extension != 'xlsx'):
            return redirect('/')
        pass

    def readRows(self):
        self.arrayData = []
        rows = self.ws.iter_rows(2)
        # print(rows)
        for row in rows:
            data = {}
            for title, cell in zip(self.headers, row):
                data[title] = cell.value
            self.arrayData.append(data)
        # print(self.arrayData)

    def getHeaders(self):
        self.oldHeaders = [cell.value for cell in next(self.ws.rows)]
        return len(self.oldHeaders)

    def loadFile(self):
        wb = load_workbook(self.file)
        # open file
        self.ws = wb.active
        # get old headers of file
        self.setHeaders()
        self.readRows()

    def deleteColumn(self):
        # print (self.dColumns)
        for item in self.dColumns:
            # print(item)
            index = self.oldHeaders.index(item)
            # print(index)
            self.ws.delete_cols(index+1)
            # print(self.ws)
            self.oldHeaders.remove(item)
        self.readRows()

    def setHeaders(self):
        count = self.getHeaders()
        print(count)
        if (request.form.get('newHeadersCB') == 'on' and count == len(self.newHeaders)):
            # print('new headres')
            self.headers = self.newHeaders
        else:
            self.headers = self.oldHeaders
        self.readRows()
        return True

    def saveCsv(self):
        print(self.arrayData)
        with open('out.csv', 'w', newline='', encoding="utf-8") as output_file:
            dict_writer = csv.DictWriter(output_file,  self.headers)
            dict_writer.writeheader()
            # dict_writer.writerows(self.arrayData)
