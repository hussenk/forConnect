import csv
import helpers
from flask import jsonify, make_response, request, send_file, send_from_directory
from openpyxl import load_workbook


class serviceAPI:

    def run(self):
        pass
        self.handelRequest()
        self.handelForm()
        self.handelFile()
        self.readHeaders()
        self.deleteColumn()
        self.setHeaders()
        self.getNextColumn()
        self.readRows()
        self.replaceText()
        self.createCSV()

    def __init__(self, path):
        self.statusCode = 200
        self.errors = []
        self.messages = []
        self.arrayData = []
        self.newHeaders = []
        self.delete_columns = []
        self.setNext = False
        self.path = path

    def handelRequest(self):

        if request.method == 'GET':
            self.statusCode = 402  # method not allowed
            return self.errors.append('Request Get')

        if ('upload' not in request.files):
            self.statusCode = 422
            return self.errors.append('No File')

        extension = ''
        if (len(request.files['upload'].filename) <= 0):
            self.statusCode = 422
            return self.errors.append('check File')

        file = request.files['upload']
        name, extension = file.filename.rsplit('.', 1)
        if(extension != 'xlsx'):
            self.statusCode = 422
            return self.errors.append('Type of File')

        return self.messages.append('handelRequest done')

    def handelFile(self):
        self.wb = load_workbook(self.file)
        self.ws = self.wb.active
        return self.messages.append('handelFile done')

    def readHeaders(self):
        self.headers = [cell.value for cell in next(self.ws.rows)]
        return self.messages.append('readHeaders done')

    def handelForm(self):

        if (request.form.get('is_new_headers_on') == '1'):
            self.newHeaders = helpers.stringSpelter(
                request.form.get('headers'))

        if (request.form.get('is_delete_on') == '1'):
            self.delete_columns = helpers.stringSpelter(
                request.form.get('delete_columns'))

        if (request.form.get('switch_to_next_on') == '1'):
            self.setNext = True

        if(request.form.get('searching_column')):
            self.searching_column = request.form.get('searching_column')
        else:
            self.searching_column = ''

        if(request.form.get('searching_value')):
            self.searching_value = request.form.get('searching_value')
        else:
            self.searching_value = ''

        if(request.form.get('replaceValue')):
            self.replaceValue = request.form.get('replaceValue')
        else:
            self.replaceValue = ''

        self.file = request.files['upload']
        return self.messages.append('handelForm done')

    def setHeaders(self):
        if(len(self.newHeaders)):
            print('new')
            self.headers = self.newHeaders
        return self.messages.append('setHeaders done')

    def deleteColumn(self):
        temp = self.headers.copy()
        if (len(self.delete_columns)):
            for item in self.delete_columns:
                if (item in temp):
                    index = temp.index(item)
                    self.ws.delete_cols(index+1)
                    temp.remove(item)
                else:
                    self.errors.append(
                        'error delete item not exist: ' + item)

        self.headers = temp
        return self.messages.append('deleteColumn done')

    def readRows(self):
        rows = self.ws.iter_rows(2)
        for row in rows:
            data = {}
            for title, cell in zip(self.headers, row):
                data[title] = cell.value
            self.arrayData.append(data)
        return self.messages.append('readRows done')

    def getNextColumn(self):
        if(self.searching_column in self.headers and self.setNext):
            index = self.headers.index(self.searching_column)+1
            if(self.headers.index(self.searching_column)+1 > len(self.headers)-1):
                index = 0
            self.nextKey = self.headers[index]
        return self.messages.append('getNextColumn done')

    def replaceText(self):

        if (len(self.searching_column) <= 0):
            return
        if (self.searching_column not in self.headers):
            return self.errors.append('the column dose not exist or his been deleted ' + self.searching_column)

        self.findInRow = []
        row = 1
        for i in self.arrayData:
            row += 1
            if(i[self.searching_column].find(self.searching_value) >= 0):
                if(len(self.replaceValue) == 0 and self.setNext):
                    i.update(
                        {self.searching_column: i[self.nextKey], self.nextKey: i[self.searching_column]})
                elif(self.setNext):
                    i.update(
                        {self.searching_column: self.replaceValue, self.nextKey: i[self.searching_column]})
                else:
                    i.update(
                        {self.searching_column: i[self.searching_column].replace(self.searching_value, self.replaceValue)})
                self.messages.append('you Can find it in row: ' + str(row))
                self.findInRow.append(row)
        return self.messages.append('replaceText done')

    def createCSV(self):
        with open(self.path+'\out.csv', 'w', newline='', encoding="utf-8") as output_file:
            dict_writer = csv.DictWriter(output_file,  self.headers)
            dict_writer.writeheader()
            dict_writer.writerows(self.arrayData)

        return self.messages.append('createCSV done')

    def response(self):

        # print('message ', self.messages)
        # print('errors ', self.errors)
        return make_response(jsonify(
            {
                'message': self.messages,
                'errors': self.errors,
            }
        ), self.statusCode)
