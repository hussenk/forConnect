import csv
import helpers
from flask import Flask, render_template, request, send_file, send_from_directory, url_for, redirect,flash
from openpyxl import load_workbook
import sys


class serviceV2:

    def __init__(self):
        self.arrayData = []
        self.newHeaders = []
        self.dColumns = []
        self.setNext = False

    # handel the request if "False" then something wrong
    def handelRequest(self):
        if request.method == 'GET':
            print('request get') # TODO error flash message
            flash('Request Get','error')
            return False

        if ('upload' not in request.files):
            flash('No File','error')
            print('no file')  # TODO error flash message
            return False
        file = request.files['upload']
        name, extension = file.filename.rsplit('.', 1)
        if(extension != 'xlsx'):
            flash('Type of File','error')
            print('Type of File')  # TODO error flash message
            return False

        return True

    def handelFile(self):
        self.wb = load_workbook(self.file)
        self.ws = self.wb.active
        return True

    def readHeaders(self):
        self.headers = [cell.value for cell in next(self.ws.rows)]
        return True

    def handelForm(self):
        if (request.form.get('newHeadersCB') == 'on'):
            self.newHeaders = helpers.stringSpelter(request.form['newHeaders'])
        if (request.form.get('dColumnsCB') == 'on'):
            self.dColumns = helpers.stringSpelter(request.form['dColumns'])
        if (request.form.get('valueNextCB') == 'on'):
            self.setNext = True
        self.column = request.form['column']
        self.value = request.form['value']
        self.replaceValue = request.form['replaceValue']

        self.file = request.files['upload']
        return True

    def setHeaders(self):
        if(len(self.newHeaders)):
            print('new')
            self.headers = self.newHeaders
        return True

    def deleteColumn(self):
        temp = self.headers.copy()
        if (len(self.dColumns)):
            for item in self.dColumns:
                if (item in temp):
                    index = temp.index(item)
                    self.ws.delete_cols(index+1)
                    temp.remove(item)
                else:
                    flash('error delete item not exist: \t' + item,'error')
                    print('error delete item not exist: \t' + item) # TODO error flash message


        self.headers = temp
        return True

    def readRows(self):
        rows = self.ws.iter_rows(2)
        for row in rows:
            data = {}
            for title, cell in zip(self.headers, row):
                data[title] = cell.value
            self.arrayData.append(data)
        return True

    def getNextColumn(self):
        if(self.column in self.headers and self.setNext):
            index = self.headers.index(self.column)+1
            if(self.headers.index(self.column)+1 > len(self.headers)-1):
                index = 0
            self.nextKey = self.headers[index]
        return True

    def replaceText(self):
        if (len(self.column) <= 0):
            return False
        self.findInRow = []

        row = 1
        for i in self.arrayData:
            row += 1
            if(i[self.column].find(self.value) >= 0):
                if(self.setNext):
                    i.update(
                        {self.column: self.replaceValue, self.nextKey: i[self.column]})
                else:
                    i.update(
                        {self.column: i[self.column].replace(self.value, self.replaceValue)})
                self.findInRow.append(row)
        return True

    def createCSV(self, path):
        with open(path+'\out.csv', 'w', newline='', encoding="utf-8") as output_file:
            dict_writer = csv.DictWriter(output_file,  self.headers)
            dict_writer.writeheader()
            dict_writer.writerows(self.arrayData)
